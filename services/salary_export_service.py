from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def generate_salary_excel(salary_data):
    wb = Workbook()

    # =============================
    # SHEET 1 — EMPLOYEE SALARIES
    # =============================
    ws = wb.active
    ws.title = "Employee Salaries"

    headers = [
        "Employee",
        "Username",
        "Role",
        "Period Start",
        "Period End",
        "Base Salary",
        "Shipment Bonus",
        "Bonus",
        "Gross Salary",
        "Tax Percent",
        "Tax Amount",
        "Net Salary",
        "Created At"
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    total_gross = 0.0
    total_tax = 0.0
    total_net = 0.0

    top_employee = ""
    top_salary = 0.0

    for row in salary_data:
        base_salary = float(row.get("base_salary", 0) or 0)
        shipment_bonus = float(row.get("shipment_bonus", 0) or 0)
        bonus = float(row.get("bonus", 0) or 0)
        tax_percent = float(row.get("tax_percent", 0) or 0)
        net_salary = float(row.get("total_salary", 0) or 0)

        gross_salary = base_salary + shipment_bonus + bonus
        tax_amount = gross_salary * tax_percent / 100

        employee_name = row.get("staff_full_name") or row.get("employee") or ""
        username = row.get("staff_username") or row.get("username") or ""
        job_title = row.get("job_title") or row.get("role") or ""

        ws.append([
            employee_name,
            username,
            job_title,
            str(row.get("period_start", "")),
            str(row.get("period_end", "")),
            round(base_salary, 2),
            round(shipment_bonus, 2),
            round(bonus, 2),
            round(gross_salary, 2),
            round(tax_percent, 2),
            round(tax_amount, 2),
            round(net_salary, 2),
            str(row.get("created_at", ""))
        ])

        total_gross += gross_salary
        total_tax += tax_amount
        total_net += net_salary

        if net_salary > top_salary:
            top_salary = net_salary
            top_employee = employee_name

    ws.append([])
    ws.append([
        "TOTAL",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        round(total_gross, 2),
        "",
        round(total_tax, 2),
        round(total_net, 2),
        ""
    ])

    total_row_index = ws.max_row
    for cell in ws[total_row_index]:
        cell.font = Font(bold=True)

    desired_widths = {
        1: 24,
        2: 18,
        3: 16,
        4: 14,
        5: 14,
        6: 14,
        7: 16,
        8: 12,
        9: 14,
        10: 12,
        11: 12,
        12: 14,
        13: 22
    }

    for col_idx, width in desired_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # =============================
    # SHEET 2 — COMPANY SUMMARY
    # =============================
    ws2 = wb.create_sheet(title="Company Summary")

    ws2.append(["Metric", "Value"])
    ws2.append(["Total Gross Payroll", round(total_gross, 2)])
    ws2.append(["Total Tax", round(total_tax, 2)])
    ws2.append(["Total Net Salaries", round(total_net, 2)])
    ws2.append(["Top Employee", top_employee])
    ws2.append(["Top Net Salary", round(top_salary, 2)])

    for cell in ws2[1]:
        cell.font = Font(bold=True)

    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 20

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return stream